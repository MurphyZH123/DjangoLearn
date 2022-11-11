#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
__title__  = openpyxl操作Excel工具类
"""

import openpyxl


class ExcelUtil:
    workBook = None
    workSheet = None

    def load_excel(self, path):
        """
        加载Excel
        :param path: 需要打开的Excel的路径
        """
        self.workBook = openpyxl.load_workbook(path)

    def get_sheet_by_name(self, name):
        """
        获取sheet对象
        :param name: sheet名
        """
        self.workSheet = self.workBook[name]

    def get_sheet_by_index(self, index=0):
        """
        获取sheet对象
        :param index: sheet的索引
        """
        # 获取workBook里所有的sheet名 -> list
        sheet_names = self.workBook.get_sheet_names()
        # 根据索引获取指定sheet
        self.workSheet = self.workBook[sheet_names[index]]

    def get_cell_value(self, col, row):
        """
        获取cell的值
        :param col: 所在列
        :param row: 所在行
        """
        try:
            return self.workSheet.cell(column=col, row=row).value
        except BaseException as e:
            return None

    def get_cell_value_by_xy(self, str):
        """
        获取cell的值
        :param str: 坐标
        """
        try:
            return self.workSheet[str].value
        except BaseException as e:
            return None

    def get_sheet_rows(self):
        """
        获取最大行数
        """
        return self.workSheet.max_row

    def get_sheet_cols(self):
        """
        获取最大列数
        """
        return self.workSheet.max_column

    def write_data_cell(self, row, col, value, path):
        """
        写入数据,按照行列号来添加数据
        """
        try:
            self.workSheet = self.workBook.active
            self.workSheet.cell(column=col, row=row, value=value)
            self.workBook.save(path)
        except BaseException as e:
            print(e)
            return None

    def write_data_append(self, value, path):
        """
        写入数据,底部追加,value是list,dict等
        """
        try:
            self.workSheet = self.workBook.active
            self.workSheet.append(value)
            self.workBook.save(path)
        except BaseException as e:
            print(e)
            return None

    def get_excel_data(self):
        """
        获取表所有数据
        :return: list
        """
        # 方式一
        data_list = tuple(self.workSheet.values)
        # 方式二
        # data_list = []
        # for i in range(self.get_sheet_rows()):
        #     data_list.append(self.get_row_value(i + 2))
        return data_list

    def get_row_value(self, row):
        """
        获取某一行的内容
        :param row: 第几行 -> str  **从1开始**
        :return: list
        """
        # 方式一
        row_list = self.get_excel_data()[row]
        # 方式二
        # row_list = []
        # for i in self.workSheet[str(row + 1)]:
        #     row_list.append(i.value)
        return row_list

    def get_col_value(self, col='A'):
        """
        获取某一列的内容
        :param col: 第几列 -> str
        :return: list
        """
        col_list = []
        for i in self.workSheet[col]:
            col_list.append(i.value)
        return col_list

    def get_cell_by_range(self):
        """
        获取所选范围内的值,不同列
        :param r1: 所选范围开始的单元格 -> ["A1:B3"]
        :param r2: 所选范围结束的单元格 -> ["A1:B3"]
        """
        range_list = []
        cells = self.workSheet['A1:B3']
        for c1, c2 in cells:
            range_list.append((c1.value, c2.value))
        return range_list

    def get_cell_by_inner(self, min_row=1, max_row=1, min_col=1, max_col=1):
        """
        选择某个区间的值
        :param min_row: 最小行
        :param max_row: 最大行
        :param min_col: 最小列
        :param max_col: 最大列
        """
        inner_list = []
        for row in self.workSheet.iter_rows(min_row, max_row, min_col, max_col):
            for cell in row:
                inner_list.append(cell.value)
        return inner_list

    def del_cell_by_col(self, path,col=2):
        """
        :param col:要删除的列
        :param path:文件保存路径
        """
        # self.workSheet = self.workBook.active
        try:
            self.workSheet = self.workBook.active
            self.workSheet.delete_cols(col)
            self.workBook.save(path)
        except BaseException as e:
            print(e)
            return None

    def del_cell_by_row(self,path,row=1):
        """
        删除指定的行
        :param path: 文件保存路径
        :param row: 要删除的行
        :return:
        """
        try:
            self.workSheet = self.workBook.active
            self.workSheet.rows(row)
            self.workBook.save(path)
        except BaseException as e:
            print(e)
            return None

    def get_row_num(self, case_id):
        """
        获取行号
        :param case_id: 用例编号
        :return:
        """
        num = 1
        col_data = self.get_col_value()
        for data in col_data:
            if case_id == data:
                return num
            num += 1
        return 0

    def edit_cell_by_format(self, path):
        """
        将单元格修改成文本类型
        """
        # try:
        #     self.workSheet = self.workBook.active
        #     col = self.workSheet.column_dimensions['C']
        #     col.number_format = '@'
        #     self.workBook.save(path)
        # except BaseException as e:
        #     print(e)
        #     return None
        try:
            self.workSheet = self.workBook.active
            self.workSheet.cell(2, 3).number_format = '@'
            self.workBook.save(path)
        except BaseException as e:
            print(e)
            return None


excelUtil = ExcelUtil()

if __name__ == '__main__':
    path = '/Users/00422829/Documents/want_club_test/test_data/want_backend/coupons_code001.xlsx'
    # 读取excel文件
    excelUtil.load_excel(path)
    # 获取某个sheet
    excelUtil.get_sheet_by_name("工作表1")
    # excelUtil.get_sheet_by_index()

    # 删除某一行
    # excelUtil.del_cell_by_col(path,2)
    # data = excelUtil.get_excel_data()
    # print(data)
    # # 获取某个cell的值
    # data = excelUtil.get_cell_value(col=1, row=1)
    # print(data)
    # data = excelUtil.get_cell_value_by_xy("A3")
    # print(data)

    # # 获取sheet行数
    # data = excelUtil.get_sheet_rows()
    # print(data)

    # # 获取sheet列数
    # data = excelUtil.get_sheet_cols()
    # print(data)

    # # 获取某一行数据
    # data = excelUtil.get_row_value(1)
    # print(data)

    # # 获取某一列数据
    # data = excelUtil.get_col_value('A')
    # print(data)

    # # 写入数据
    excelUtil.write_data_cell(row=3, col=2, value=2.00, path=path)

    # # 获取全部数据
    # data = excelUtil.get_excel_data()
    # print(data)

    # # 获取行号
    # data = excelUtil.get_row_num('imooc_001')
    # print(data)

    # 读取指定区域的值
    # data = excelUtil.get_cell_by_inner(1,5,1,1)
    # print(data)

    # 修改单元格格式
    excelUtil.edit_cell_by_format(path)
